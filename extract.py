import pandas as pd
from bcb import sgs
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
import sys
import time
import os
import shutil
from typing import List, Optional

# --- Global Configurations ---
START_DATE = '2010-01-01'

# Dynamic Path Definition (Portability)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, 'input_series.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, 'Resultado_BCB.xlsx')
BACKUP_FILE = os.path.join(BASE_DIR, 'Resultado_BCB_BACKUP.xlsx')

BATCH_SIZE = 10  # Batch size for API requests

def generate_column_range(max_col_str: str) -> List[str]:
    """
    Generates a list of Excel column letters starting from 'B' up to the specified column.
    
    Args:
        max_col_str: The maximum column letter (e.g., 'Z', 'AA').
        
    Returns:
        A list of column letters ['B', 'C', ..., max_col_str].
    """
    try:
        max_idx = column_index_from_string(max_col_str)
        return [get_column_letter(i) for i in range(2, max_idx + 1)]
    except Exception:
        return []

from concurrent.futures import ThreadPoolExecutor, as_completed

MAX_WORKERS = 5  # Adjust based on system/network limits

def process_batch_chunk(chunk: List[int], start_date: str) -> List[pd.DataFrame]:
    """
    Helper function to process a single batch of series codes.
    """
    chunk_results = []
    try:
        # Batch download attempt (Performance Optimization)
        # Note: Printing in threads might interleave, but acceptable for simple logs
        print(f"  > Processing batch (series {chunk[0]}...)...")
        df_chunk = sgs.get(chunk, start=start_date)
        chunk_results.append(df_chunk)
        
    except Exception as e:
        print(f"  ! FAILURE in batch ({chunk[0]}...). Starting individual recovery mode...")
        # Fallback: Process individually
        for code in chunk:
            try:
                # Attempt 1: Strict download with start date
                ts = sgs.get(code, start=start_date)
                ts.name = code
                chunk_results.append(ts)
            except Exception:
                try:
                    # Attempt 2: Full history download and local filtering
                    # Useful for series starting AFTER start_date
                    ts = sgs.get(code)
                    if not ts.empty:
                        ts = ts[ts.index >= start_date]
                        if not ts.empty:
                            ts.name = code
                            chunk_results.append(ts)
                        else:
                            print(f"    - Series {code}: Data retrieved but none within period.")
                except:
                    print(f"    - Series {code} failed permanently.")
    return chunk_results

def download_series_batch(series_codes: List[int], start_date: str) -> pd.DataFrame:
    """
    Downloads BCB series in parallel batches with fallback to individual handling on failure.
    
    Args:
        series_codes: List of series codes (integers).
        start_date: Start date string in 'YYYY-MM-DD' format.
        
    Returns:
        DataFrame containing the consolidated data.
    """
    results: List[pd.DataFrame] = []
    
    # Remove duplicates and ensure integer typing
    series_codes = sorted(list(set(series_codes)))
    total = len(series_codes)
    
    print(f"Starting parallel download of {total} series with {MAX_WORKERS} workers...")

    # Create chunks
    chunks = [series_codes[i : i + BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Submit all tasks
        future_to_chunk = {executor.submit(process_batch_chunk, chunk, start_date): chunk for chunk in chunks}
        
        # Process results as they complete
        for future in as_completed(future_to_chunk):
            try:
                chunk_data = future.result()
                results.extend(chunk_data)
            except Exception as e:
                print(f"CRITICAL WORKER ERROR: {e}")

    # Data Consolidation
    if results:
        print("Consolidating data...", end=" ")
        df_final = pd.concat(results, axis=1)
        # Standardize columns as integers for optimized indexing
        df_final.columns = pd.to_numeric(df_final.columns, errors='coerce').fillna(0).astype(int)
        print("OK")
        return df_final
    else:
        return pd.DataFrame()

def main():
    print("=== Extract.py: Batch Extraction Process Started ===")
    
    start_time = datetime.datetime.now()
    print(f"Execution Start: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Base Directory: {BASE_DIR}")
    
    exit_code = 0
    
    # --- Automatic Safety Backup ---
    if os.path.exists(OUTPUT_FILE):
        try:
            print(f"Existing output file detected. Creating backup at: {BACKUP_FILE} ...", end=" ")
            shutil.copy2(OUTPUT_FILE, BACKUP_FILE)
            print("SUCCESS.")
        except IOError as e:
            print(f"\narting warning: Backup creation failed ({e}). Execution will continue.")
        except Exception as e:
            print(f"\nWARNING: Unexpected error creating backup ({e}). Execution will continue.")
    else:
        print("Output file not found. Backup skipped (first run or file removed).")

    try:
        # 1. Master Index Generation (Monthly Frequency)
        today = datetime.datetime.today().strftime('%Y-%m-%d')
        master_index = pd.date_range(start=START_DATE, end=today, freq='MS')
        master_index.name = 'Data'

        # 2. Input File Reading
        try:
            print(f"Reading configuration file: {INPUT_FILE}")
            df_input = pd.read_excel(INPUT_FILE, usecols=['Codigo', 'Coluna', 'Aba'])
            
            # Data Cleaning and Processing
            df_input = df_input.dropna(subset=['Codigo', 'Aba'])
            df_input['Aba'] = df_input['Aba'].astype(str).str.strip()
            df_input['Coluna'] = df_input['Coluna'].astype(str).str.strip().str.upper()
            
            # Series Code Sanitization
            df_input['Codigo_Clean'] = pd.to_numeric(df_input['Codigo'], errors='coerce').fillna(0).astype(int)
            df_input = df_input[df_input['Codigo_Clean'] > 0]
            
        except FileNotFoundError:
            print(f"CRITICAL ERROR: Input file not found at: {INPUT_FILE}")
            exit_code = 1
            return
        except Exception as e:
            print(f"CRITICAL ERROR reading input file: {e}")
            exit_code = 1
            return

        # 3. Centralized Download Execution (Batch Processing)
        unique_codes = df_input['Codigo_Clean'].unique().tolist()
        
        if not unique_codes:
            print("No valid series codes found.")
            exit_code = 1
            return

        df_global_data = download_series_batch(unique_codes, START_DATE)
        
        # Reindexing to Master Index
        if not df_global_data.empty:
            df_global_data = df_global_data.reindex(master_index)

        # 4. Data Distribution by Sheet
        output_dfs = {}
        unique_sheets = df_input['Aba'].unique()

        print(f"\nDistributing data to {len(unique_sheets)} sheets...")

        for sheet in unique_sheets:
            mask = df_input['Aba'] == sheet
            df_config = df_input[mask]
            
            # Sheet Column Structure Definition
            try:
                col_indices = df_config['Coluna'].apply(column_index_from_string)
                max_col_str = get_column_letter(col_indices.max())
                target_cols = generate_column_range(max_col_str)
            except:
                continue
                
            # Sheet DataFrame Initialization
            df_sheet = pd.DataFrame(index=master_index, columns=target_cols)
            
            # Data Mapping
            for _, row in df_config.iterrows():
                col_dest = row['Coluna']
                cod_orig = row['Codigo_Clean']
                
                # Data Transfer
                if cod_orig in df_global_data.columns:
                    df_sheet[col_dest] = df_global_data[cod_orig]
            
            output_dfs[sheet] = df_sheet

        # 5. Final File Export
        if output_dfs:
            try:
                print(f"Saving output file: {OUTPUT_FILE} ...")
                with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
                    for sheet, df in output_dfs.items():
                        df.to_excel(writer, sheet_name=sheet)
                print("Process completed successfully.")
            except PermissionError:
                print(f"CRITICAL ERROR: Permission denied saving {OUTPUT_FILE}. Close the file if open.")
                exit_code = 1
            except Exception as e:
                print(f"CRITICAL ERROR saving Excel file: {e}")
                exit_code = 1
        else:
            print("No data generated for export.")
            exit_code = 1

    except Exception as e:
        print(f"UNHANDLED ERROR during execution: {e}")
        exit_code = 1

    finally:
        # End Time Marker
        end_time = datetime.datetime.now()
        duration = end_time - start_time
        print(f"\nExecution End: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Total processing time: {duration}")
        sys.exit(exit_code)

if __name__ == "__main__":
    main()